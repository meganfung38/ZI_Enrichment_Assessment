from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import io
import pandas as pd

class ExcelService:
    """Service for exporting lead analysis data to Excel format"""
    
    def __init__(self):
        # RingCentral Brand Colors
        self.rc_cerulean = "0684BC"      # RingCentral primary blue
        self.rc_orange = "FF7A00"        # RingCentral orange
        self.rc_ocean = "002855"         # RingCentral dark blue
        self.rc_linen = "F1EFEC"         # RingCentral background
        self.rc_ash = "C8C2B4"           # RingCentral light gray
        self.rc_warm_black = "2B2926"    # RingCentral dark gray
        
        # Excel Styling with RingCentral Colors
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color=self.rc_cerulean, end_color=self.rc_cerulean, fill_type="solid")
        self.summary_font = Font(bold=True, color=self.rc_ocean)
        self.summary_fill = PatternFill(start_color=self.rc_linen, end_color=self.rc_linen, fill_type="solid")
        self.title_font = Font(bold=True, size=16, color=self.rc_ocean)
        self.accent_fill = PatternFill(start_color=self.rc_orange, end_color=self.rc_orange, fill_type="solid")
        self.border = Border(
            left=Side(style='thin', color=self.rc_ash),
            right=Side(style='thin', color=self.rc_ash),
            top=Side(style='thin', color=self.rc_ash),
            bottom=Side(style='thin', color=self.rc_ash)
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def create_lead_analysis_excel(self, analysis_data, summary_data, query_info=None, filename_prefix="lead_analysis"):
        """Create Excel file from lead analysis data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Lead Analysis"
        
        # Add title and metadata
        current_row = 1
        ws.merge_cells(f'A{current_row}:R{current_row}')
        ws[f'A{current_row}'] = "ZoomInfo Lead Quality Analysis Report"
        ws[f'A{current_row}'].font = self.title_font
        ws[f'A{current_row}'].alignment = self.center_alignment
        current_row += 1
        
        # Add generation timestamp
        ws.merge_cells(f'A{current_row}:R{current_row}')
        ws[f'A{current_row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{current_row}'].alignment = self.center_alignment
        current_row += 2
        
        # Add summary section if provided
        if summary_data:
            current_row = self._add_summary_section(ws, summary_data, query_info, current_row)
            current_row += 1
        
        # Add headers for all lead data fields plus assessment outputs
        headers = [
            "Lead ID", "Email", "First Channel", "Segment Name", "Company Size Range",
            "Website", "Company", "ZI Website", "ZI Company Name", "ZI Employees", "Email Domain",
            "Not in TAM", "Suspicious Enrichment", "Confidence Score", 
            "Explanation", "Corrections", "Inferences", "AI Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        current_row += 1
        
        # Add data rows
        for lead in analysis_data:
            self._add_lead_row(ws, lead, current_row)
            current_row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Create file buffer
        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        
        return file_buffer, filename
    
    def _add_summary_section(self, ws, summary_data, query_info, start_row):
        """Add summary statistics section to the worksheet"""
        ws.merge_cells(f'A{start_row}:D{start_row}')
        ws[f'A{start_row}'] = "Analysis Summary"
        ws[f'A{start_row}'].font = self.summary_font
        ws[f'A{start_row}'].fill = self.summary_fill
        start_row += 1
        
        # Summary statistics
        summary_items = [
            ("Total Leads Analyzed", summary_data.get('leads_analyzed', 0)),
            ("Leads with Issues", summary_data.get('leads_with_issues', 0)),
            ("Issue Percentage", f"{summary_data.get('issue_percentage', 0)}%"),
            ("Average Confidence Score", summary_data.get('avg_confidence_score', 0)),
            ("Not in TAM Count", summary_data.get('not_in_tam_count', 0)),
            ("Suspicious Enrichment Count", summary_data.get('suspicious_enrichment_count', 0)),
        ]
        
        if summary_data.get('total_query_results'):
            summary_items.insert(0, ("Total Query Results", summary_data.get('total_query_results', 0)))
        
        if summary_data.get('ai_assessments_successful') is not None:
            summary_items.extend([
                ("AI Assessments Successful", summary_data.get('ai_assessments_successful', 0)),
                ("AI Assessments Failed", summary_data.get('ai_assessments_failed', 0))
            ])
        
        for label, value in summary_items:
            ws[f'A{start_row}'] = label
            ws[f'B{start_row}'] = value
            ws[f'A{start_row}'].font = Font(bold=True)
            start_row += 1
        
        # Query info if available
        if query_info:
            start_row += 1
            ws.merge_cells(f'A{start_row}:D{start_row}')
            ws[f'A{start_row}'] = "Query Information"
            ws[f'A{start_row}'].font = self.summary_font
            ws[f'A{start_row}'].fill = self.summary_fill
            start_row += 1
            
            if query_info.get('original_query'):
                ws[f'A{start_row}'] = "Original Query"
                ws[f'B{start_row}'] = query_info['original_query']
                ws[f'B{start_row}'].alignment = self.wrap_alignment
                start_row += 1
            
            if query_info.get('execution_time'):
                ws[f'A{start_row}'] = "Execution Time"
                ws[f'B{start_row}'] = query_info['execution_time']
                start_row += 1
        
        return start_row + 1
    
    def _add_lead_row(self, ws, lead, row):
        """Add a single lead's data to the worksheet"""
        # Extract confidence assessment data
        confidence_assessment = lead.get('confidence_assessment', {})
        confidence_score = confidence_assessment.get('confidence_score', '') if confidence_assessment else ''
        
        # Format explanation bullets
        explanation_bullets = confidence_assessment.get('explanation_bullets', []) if confidence_assessment else []
        explanation_text = '\n'.join(explanation_bullets) if explanation_bullets else ''
        
        # Format corrections and inferences
        corrections = confidence_assessment.get('corrections', {}) if confidence_assessment else {}
        corrections_text = json.dumps(corrections, indent=2) if corrections else ''
        
        inferences = confidence_assessment.get('inferences', {}) if confidence_assessment else {}
        inferences_text = json.dumps(inferences, indent=2) if inferences else ''
        
        # Data for each column (matching header order)
        row_data = [
            lead.get('Id', ''),                          # Lead ID
            lead.get('Email', ''),                       # Email
            lead.get('First_Channel__c', ''),            # First Channel
            lead.get('SegmentName', ''),                 # Segment Name
            lead.get('LS_Company_Size_Range__c', ''),    # Company Size Range
            lead.get('Website', ''),                     # Website
            lead.get('Company', ''),                     # Company
            lead.get('ZI_Website__c', ''),               # ZI Website
            lead.get('ZI_Company_Name__c', ''),          # ZI Company Name
            lead.get('ZI_Employees__c', ''),             # ZI Employees
            lead.get('email_domain', ''),                # Email Domain
            'Yes' if lead.get('not_in_TAM') else 'No',   # Not in TAM
            'Yes' if lead.get('suspicious_enrichment') else 'No',  # Suspicious Enrichment
            confidence_score,                            # Confidence Score
            explanation_text,                            # Explanation
            corrections_text,                            # Corrections
            inferences_text,                             # Inferences
            lead.get('ai_assessment_status', '')         # AI Status
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.border
            
            # Special formatting for certain columns with RingCentral colors
            if col in [12, 13]:  # Boolean flags (Not in TAM, Suspicious Enrichment)
                cell.alignment = self.center_alignment
                if value == 'Yes':
                    # Light red background for issues (softer than pure red)
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    cell.font = Font(bold=True, color=self.rc_warm_black)
            elif col == 14:  # Confidence score with RingCentral color scheme
                cell.alignment = self.center_alignment
                cell.font = Font(bold=True, color="FFFFFF")
                if isinstance(value, (int, float)) and value > 0:
                    if value >= 80:
                        # High confidence: RingCentral Cerulean (primary blue)
                        cell.fill = PatternFill(start_color=self.rc_cerulean, end_color=self.rc_cerulean, fill_type="solid")
                    elif value >= 60:
                        # Medium confidence: RingCentral Orange
                        cell.fill = PatternFill(start_color=self.rc_orange, end_color=self.rc_orange, fill_type="solid")
                    else:
                        # Low confidence: Muted red with dark text
                        cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            elif col in [15, 16, 17]:  # Text fields that might be long (Explanation, Corrections, Inferences)
                cell.alignment = self.wrap_alignment
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        column_widths = {
            1: 20,   # Lead ID
            2: 25,   # Email
            3: 15,   # First Channel
            4: 15,   # Segment Name
            5: 18,   # Company Size Range
            6: 20,   # Website
            7: 20,   # Company
            8: 20,   # ZI Website
            9: 20,   # ZI Company Name
            10: 12,  # ZI Employees
            11: 15,  # Email Domain
            12: 12,  # Not in TAM
            13: 15,  # Suspicious Enrichment
            14: 12,  # Confidence Score
            15: 40,  # Explanation
            16: 25,  # Corrections
            17: 25,  # Inferences
            18: 12   # AI Status
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Set row heights for better readability
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 25
    
    def create_single_lead_excel(self, lead_data, filename_prefix="lead_confidence"):
        """Create Excel file for a single lead confidence assessment"""
        return self.create_lead_analysis_excel(
            analysis_data=[lead_data],
            summary_data=None,
            query_info=None,
            filename_prefix=filename_prefix
        )
    
    def parse_excel_file(self, file_content):
        """Parse uploaded Excel file and return sheet names and preview data"""
        try:
            # Load workbook from file content
            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            sheet_names = wb.sheetnames
            
            # Get preview data from first sheet
            first_sheet = wb[sheet_names[0]]
            
            # Read first 10 rows for preview
            preview_data = []
            headers = []
            
            for row_idx, row in enumerate(first_sheet.iter_rows(values_only=True)):
                if row_idx == 0:
                    # First row as headers
                    headers = [cell if cell is not None else f"Column_{i+1}" for i, cell in enumerate(row)]
                elif row_idx < 11:  # First 10 data rows
                    row_data = [cell if cell is not None else "" for cell in row]
                    # Pad row to match header length
                    while len(row_data) < len(headers):
                        row_data.append("")
                    preview_data.append(row_data[:len(headers)])  # Trim to header length
                else:
                    break
            
            wb.close()
            
            # Calculate total rows safely (max_row can be None for empty sheets)
            total_rows = (first_sheet.max_row - 1) if first_sheet.max_row else 0
            
            return {
                'success': True,
                'sheet_names': sheet_names,
                'headers': headers,
                'preview_data': preview_data,
                'total_rows': max(total_rows, len(preview_data))  # Use actual data count if max_row is unreliable
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error parsing Excel file: {str(e)}"
            }
    
    def extract_lead_ids_from_excel(self, file_content, sheet_name, lead_id_column):
        """Extract Lead IDs from specified column in Excel file"""
        try:
            # Use pandas for easier data extraction - read as string to preserve Lead ID format
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_name, dtype={lead_id_column: str})
            
            if lead_id_column not in df.columns:
                return {
                    'success': False,
                    'error': f"Column '{lead_id_column}' not found in sheet '{sheet_name}'"
                }
            
            # Extract Lead IDs and remove null/empty values
            lead_ids = df[lead_id_column].dropna().astype(str).tolist()
            # Remove empty strings and whitespace-only strings, and handle Excel formatting issues
            cleaned_lead_ids = []
            for lid in lead_ids:
                lid_str = str(lid).strip()
                # Remove any Excel formatting artifacts
                if lid_str and lid_str.lower() not in ['nan', 'none', 'null']:
                    # Handle potential floating point conversion (e.g., "1.23456789012345e+17")
                    if 'e+' in lid_str.lower():
                        try:
                            # Convert scientific notation back to full number
                            lid_str = f"{float(lid_str):.0f}"
                        except:
                            pass
                    cleaned_lead_ids.append(lid_str)
            
            lead_ids = cleaned_lead_ids
            
            # Get original data for later merging - handle NaN values
            # Replace NaN with empty string to avoid JSON serialization issues
            df_clean = df.where(pd.notnull(df), '')  # Replace NaN with empty string
            original_data = df_clean.to_dict('records')
            
            return {
                'success': True,
                'lead_ids': lead_ids,
                'original_data': original_data,
                'total_rows': len(df)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error extracting Lead IDs: {str(e)}"
            }
    
    def _convert_15_to_18_char_id(self, id_15):
        """Convert 15-character Salesforce ID to 18-character format"""
        if len(id_15) != 15:
            return id_15
        
        # Salesforce ID conversion algorithm
        suffix = ""
        for i in range(3):
            chunk = id_15[i*5:(i+1)*5]
            chunk_value = 0
            for j, char in enumerate(chunk):
                if char.isupper():
                    chunk_value += 2 ** j
            
            # Convert to base-32 character
            if chunk_value < 26:
                suffix += chr(ord('A') + chunk_value)
            else:
                suffix += str(chunk_value - 26)
        
        return id_15 + suffix

    def create_excel_with_analysis(self, original_data, analysis_results, lead_id_column, filename_prefix="excel_analysis"):
        """Create Excel file combining original data with AI analysis results"""
        try:
            # Convert original data to DataFrame and handle NaN values
            df_original = pd.DataFrame(original_data)
            # Ensure no NaN values that could cause JSON serialization issues
            df_original = df_original.where(pd.notnull(df_original), '')
            
            # Create analysis dictionary for fast lookup with both 15 and 18 char Lead IDs
            analysis_dict = {}
            print(f"🔍 DEBUG: Creating analysis dictionary from {len(analysis_results)} results")
            for i, result in enumerate(analysis_results):
                lead_id = result.get('Id', '')
                if lead_id:
                    lead_id_str = str(lead_id).strip()
                    print(f"   Result {i+1}: Lead ID '{lead_id_str}' (length: {len(lead_id_str)})")
                    
                    # Store with the original ID format
                    analysis_dict[lead_id_str] = result
                    analysis_dict[lead_id_str.upper()] = result
                    analysis_dict[lead_id_str.lower()] = result
                    
                    # If it's 18 characters, also store the 15-character version
                    if len(lead_id_str) == 18:
                        lead_id_15 = lead_id_str[:15]
                        analysis_dict[lead_id_15] = result
                        analysis_dict[lead_id_15.upper()] = result
                        analysis_dict[lead_id_15.lower()] = result
                        print(f"      Also stored as 15-char: '{lead_id_15}'")
                    
                    # If it's 15 characters, also store the 18-character version
                    elif len(lead_id_str) == 15:
                        lead_id_18 = self._convert_15_to_18_char_id(lead_id_str)
                        analysis_dict[lead_id_18] = result
                        analysis_dict[lead_id_18.upper()] = result
                        analysis_dict[lead_id_18.lower()] = result
                        print(f"      Also stored as 18-char: '{lead_id_18}'")
            
            print(f"🔍 DEBUG: Analysis dictionary contains {len(analysis_dict)} total entries")
            
            # 🔍 DEBUG: Show what Lead IDs are available in analysis dictionary
            print(f"🔍 DEBUG: Analysis dictionary contains {len(analysis_dict)} Lead ID entries:")
            unique_ids = set()
            for key in analysis_dict.keys():
                if len(key) in [15, 18]:  # Only show actual Lead IDs, not upper/lower variants
                    unique_ids.add(key)
            for i, unique_id in enumerate(sorted(unique_ids)[:5]):  # Show first 5
                print(f"   {i+1}. '{unique_id}' (length: {len(unique_id)})")
            if len(unique_ids) > 5:
                print(f"   ... and {len(unique_ids) - 5} more")
            
            # Calculate summary statistics from analysis results
            leads_analyzed = len(df_original)
            leads_with_issues = 0
            not_in_tam_count = 0
            suspicious_enrichment_count = 0
            total_confidence_score = 0
            successful_ai_assessments = 0
            
            # Add analysis columns
            analysis_columns = {
                'AI_Confidence_Score': [],
                'AI_Explanation': [],
                'AI_Corrections': [],
                'AI_Inferences': [],
                'AI_Not_in_TAM': [],
                'AI_Suspicious_Enrichment': [],
                'AI_Status': []
            }
            
            for _, row in df_original.iterrows():
                original_lead_id = str(row[lead_id_column]).strip()
                
                # Try multiple formats to find a match
                analysis = None
                for potential_id in [original_lead_id, original_lead_id.upper(), original_lead_id.lower()]:
                    if potential_id in analysis_dict:
                        analysis = analysis_dict[potential_id]
                        break
                
                if not analysis:
                    analysis = {}
                
                # 🔍 DEBUG: Log what we found for this lead
                print(f"🔍 DEBUG: Lead ID '{original_lead_id}' analysis:")
                print(f"   - Found analysis: {bool(analysis)}")
                if analysis:
                    print(f"   - Has confidence_assessment: {bool(analysis.get('confidence_assessment'))}")
                    print(f"   - AI assessment status: {analysis.get('ai_assessment_status', 'NONE')}")
                    print(f"   - not_in_TAM: {analysis.get('not_in_TAM', 'NONE')}")
                    print(f"   - suspicious_enrichment: {analysis.get('suspicious_enrichment', 'NONE')}")
                    confidence_assessment = analysis.get('confidence_assessment', {})
                    if confidence_assessment:
                        print(f"   - Confidence score: {confidence_assessment.get('confidence_score', 'NONE')}")
                        print(f"   - Explanation bullets: {len(confidence_assessment.get('explanation_bullets', []))}")
                        print(f"   - Corrections: {confidence_assessment.get('corrections', {})}")
                        print(f"   - Inferences: {confidence_assessment.get('inferences', {})}")
                
                # Extract analysis data safely
                confidence_assessment = analysis.get('confidence_assessment', {})
                
                # Handle case where confidence_assessment might be None
                if confidence_assessment is None:
                    confidence_assessment = {}
                
                # Update summary statistics
                if analysis:
                    # Count quality issues
                    if analysis.get('not_in_TAM') or analysis.get('suspicious_enrichment'):
                        leads_with_issues += 1
                    if analysis.get('not_in_TAM'):
                        not_in_tam_count += 1
                    if analysis.get('suspicious_enrichment'):
                        suspicious_enrichment_count += 1
                    
                    # Count successful AI assessments
                    if (confidence_assessment and 
                        isinstance(confidence_assessment.get('confidence_score'), (int, float)) and 
                        confidence_assessment.get('confidence_score') is not None):
                        total_confidence_score += confidence_assessment.get('confidence_score', 0)
                        successful_ai_assessments += 1
                
                # Confidence Score
                confidence_score = confidence_assessment.get('confidence_score', '') if confidence_assessment else ''
                analysis_columns['AI_Confidence_Score'].append(confidence_score)
                
                # Explanation
                explanation_bullets = confidence_assessment.get('explanation_bullets', []) if confidence_assessment else []
                explanation_text = '\n'.join(explanation_bullets) if explanation_bullets else ''
                analysis_columns['AI_Explanation'].append(explanation_text)
                
                # Corrections - handle empty dict properly
                corrections = confidence_assessment.get('corrections', {}) if confidence_assessment else {}
                corrections_text = json.dumps(corrections, indent=2) if corrections else ''
                analysis_columns['AI_Corrections'].append(corrections_text)
                
                # Inferences - handle empty dict properly  
                inferences = confidence_assessment.get('inferences', {}) if confidence_assessment else {}
                inferences_text = json.dumps(inferences, indent=2) if inferences else ''
                analysis_columns['AI_Inferences'].append(inferences_text)
                
                # Quality flags
                analysis_columns['AI_Not_in_TAM'].append(
                    'Yes' if analysis.get('not_in_TAM') else 'No'
                )
                
                analysis_columns['AI_Suspicious_Enrichment'].append(
                    'Yes' if analysis.get('suspicious_enrichment') else 'No'
                )
                
                # AI Status - improved logic
                if analysis:
                    # Check if we have a valid confidence assessment
                    has_confidence_score = (confidence_assessment and 
                                          isinstance(confidence_assessment.get('confidence_score'), (int, float)) and 
                                          confidence_assessment.get('confidence_score') is not None)
                    
                    if has_confidence_score:
                        ai_status = 'Analyzed'
                    elif analysis.get('ai_assessment_status') == 'success':
                        ai_status = 'Analyzed'
                    elif analysis.get('ai_assessment_status'):
                        ai_status = str(analysis.get('ai_assessment_status'))
                    else:
                        # If we have analysis data but no AI assessment, check if we have any AI data
                        has_ai_data = (confidence_assessment and 
                                     (confidence_assessment.get('explanation_bullets') or 
                                      confidence_assessment.get('corrections') or 
                                      confidence_assessment.get('inferences')))
                        ai_status = 'Analyzed' if has_ai_data else 'Analysis Available'
                else:
                    ai_status = 'Not Analyzed'
                
                analysis_columns['AI_Status'].append(ai_status)
            
            # Add analysis columns to original DataFrame
            for col_name, col_data in analysis_columns.items():
                df_original[col_name] = col_data
            
            # Create summary data for the summary section
            avg_confidence_score = (total_confidence_score / successful_ai_assessments) if successful_ai_assessments > 0 else 0
            summary_data = {
                'leads_analyzed': leads_analyzed,
                'leads_with_issues': leads_with_issues,
                'issue_percentage': round((leads_with_issues / leads_analyzed) * 100, 2) if leads_analyzed > 0 else 0,
                'avg_confidence_score': round(avg_confidence_score, 1),
                'not_in_tam_count': not_in_tam_count,
                'suspicious_enrichment_count': suspicious_enrichment_count,
                'ai_assessments_successful': successful_ai_assessments,
                'ai_assessments_failed': leads_analyzed - successful_ai_assessments
            }
            
            # Create Excel file with formatting
            wb = Workbook()
            ws = wb.active
            assert ws is not None  # Type assertion for linter
            ws.title = "Analysis Results"
            
            # Add title with RingCentral styling
            current_row = 1
            ws.merge_cells(f'A{current_row}:{get_column_letter(len(df_original.columns))}{current_row}')
            ws[f'A{current_row}'] = "Excel Upload Analysis Results"
            ws[f'A{current_row}'].font = self.title_font
            ws[f'A{current_row}'].alignment = self.center_alignment
            current_row += 1
            
            # Add timestamp
            ws.merge_cells(f'A{current_row}:{get_column_letter(len(df_original.columns))}{current_row}')
            ws[f'A{current_row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{current_row}'].alignment = self.center_alignment
            current_row += 2
            
            # Add summary section
            current_row = self._add_summary_section(ws, summary_data, None, current_row)
            current_row += 1
            
            # Add headers 
            header_row = current_row
            for col_idx, header in enumerate(df_original.columns, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            current_row += 1
            
            # Add data rows
            for row_idx, (_, row) in enumerate(df_original.iterrows(), current_row):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.border
                    
                    # Special formatting for AI columns with RingCentral colors
                    header = df_original.columns[col_idx - 1]
                    if header in ['AI_Not_in_TAM', 'AI_Suspicious_Enrichment']:
                        cell.alignment = self.center_alignment
                        if value == 'Yes':
                            # Light red background for issues
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                            cell.font = Font(bold=True, color=self.rc_warm_black)
                    elif header == 'AI_Confidence_Score':
                        cell.alignment = self.center_alignment
                        cell.font = Font(bold=True, color="FFFFFF")
                        if isinstance(value, (int, float)) and value > 0:
                            if value >= 80:
                                # High confidence: RingCentral Cerulean (primary blue)
                                cell.fill = PatternFill(start_color=self.rc_cerulean, end_color=self.rc_cerulean, fill_type="solid")
                            elif value >= 60:
                                # Medium confidence: RingCentral Orange
                                cell.fill = PatternFill(start_color=self.rc_orange, end_color=self.rc_orange, fill_type="solid")
                            else:
                                # Low confidence: Muted red
                                cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                    elif header in ['AI_Explanation', 'AI_Corrections', 'AI_Inferences']:
                        cell.alignment = self.wrap_alignment
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(df_original.columns, 1):
                if header.startswith('AI_'):
                    if header in ['AI_Explanation', 'AI_Corrections', 'AI_Inferences']:
                        ws.column_dimensions[get_column_letter(col_idx)].width = 40
                    elif header == 'AI_Confidence_Score':
                        ws.column_dimensions[get_column_letter(col_idx)].width = 15
                    else:
                        ws.column_dimensions[get_column_letter(col_idx)].width = 18
                else:
                    # Auto-size original columns
                    ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            # Create file buffer
            file_buffer = io.BytesIO()
            wb.save(file_buffer)
            file_buffer.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{filename_prefix}_{timestamp}.xlsx"
            
            return {
                'success': True,
                'file_buffer': file_buffer,
                'filename': filename
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error creating Excel with analysis: {str(e)}"
            } 